# src/utils/file_handler.py
# Only make minimal changes to reduce log output
import os
import pandas as pd
from datetime import datetime
import logging
import glob
from openpyxl import load_workbook
import shutil
import re
import time
import random

# Configure module-level logger with a simplified name
logger = logging.getLogger("file_handler")

def normalize_url(url):
    """
    Normalize a URL for consistent comparison.
    
    1. Convert to lowercase
    2. Remove trailing slashes
    3. Strip whitespace
    4. Sort query parameters (future enhancement)
    """
    if not url:
        return ""
    
    url = str(url).lower().strip()
    url = url.rstrip('/')
    
    # Remove URL parameters (if needed)
    # url = url.split('?')[0]
    
    return url

def read_xlsm(file_path):
    """
    Read Excel file (XLSM or XLSX) and convert to list of dictionaries
    """
    import pandas as pd
    from openpyxl import load_workbook
    import os
    
    logger.info(f"Reading: {os.path.basename(file_path)}")
    
    file_ext = os.path.splitext(file_path)[1].lower()
    
    try:
        # For XLSM files, try to use openpyxl directly to ensure we get all columns
        if file_ext == '.xlsm':
            # Load the workbook
            wb = load_workbook(filename=file_path, read_only=True, keep_vba=True)
            sheet = wb.active  # Assumes data is in the active sheet
            
            # Get column headers
            headers = [cell.value for cell in next(sheet.iter_rows())]
            
            # Read data only for non-empty rows - no limits on number of rows
            data = []
            for row in list(sheet.iter_rows())[1:]:  # Skip header only
                row_dict = {}
                # Check if row has any non-empty cells
                if any(cell.value is not None for cell in row):
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i] is not None:
                            row_dict[headers[i]] = "" if cell.value is None else cell.value
                    data.append(row_dict)
            
            logger.info(f"Read {len(data)} rows from XLSM file")
            
            # Close the workbook
            wb.close()
            return data
            
        # Fall back to pandas for all other file types
        df = pd.read_excel(file_path, engine="openpyxl")
        
        # Handle NaN values in the DataFrame
        df = df.fillna('')
        
        # No row limit - process all rows
        
        # Convert updated_at column to string if it exists
        if "updated_at" in df.columns:
            # Convert to string and remove decimal part if present
            df["updated_at"] = df["updated_at"].astype(str).apply(
                lambda x: x.split('.')[0] if '.' in x else x).str.strip()
        
        # Convert to list of dictionaries, skip empty rows
        data = []
        for _, row in df.iterrows():
            # Check if row has any non-empty values
            if any(row[col] != '' for col in df.columns):
                row_dict = {}
                for col in df.columns:
                    row_dict[col] = row[col]
                data.append(row_dict)
        
        logger.info(f"Read {len(data)} rows from Excel file")
        return data
    except Exception as e:
        logger.error(f"Error reading file: {str(e)}")
        
        try:
            # Try again with pandas
            df = pd.read_excel(file_path, engine="openpyxl")
            data = [row.to_dict() for _, row in df.iterrows()]
            logger.info(f"Read {len(data)} rows with fallback method")
            return data
        except Exception as e2:
            logger.error(f"Fallback failed: {str(e2)}")
            return []

def update_input_file(input_file, items):
    """
    Update the input file directly with the scraped data, preserving original row order,
    formatting, and file format (especially for XLSM files).
    Uses PriceLink as the key to match rows.
    """
    import pandas as pd
    import os
    from datetime import datetime
    import logging
    from openpyxl import load_workbook
    import shutil
    
    logger.info(f"Updating file: {os.path.basename(input_file)}")
    
    file_ext = os.path.splitext(input_file)[1].lower()
    
    # Get market player from first item for diagnostics
    market_player = None
    if items and len(items) > 0 and "market_player" in items[0]:
        market_player = items[0]["market_player"]
    
    # Retry parameters for file access
    max_retries = 5
    base_retry_delay = 20  # seconds
    
    # Try multiple times to update the file in case of concurrent access
    for attempt in range(1, max_retries + 1):
        try:
            # Check if the file can be opened
            if not os.path.exists(input_file):
                logger.error(f"Input file does not exist: {input_file}")
                return 0
                
            # Try to open the file to see if it's accessible
            try:
                with open(input_file, 'rb') as test_file:
                    # Just testing if we can open it
                    pass
            except PermissionError:
                if attempt < max_retries:
                    # File is being used by another process, retry after delay
                    retry_delay = base_retry_delay * (1 + random.random())  # Add some randomness
                    logger.warning(f"File is locked by another process. Retry {attempt}/{max_retries} after {retry_delay:.1f}s delay")
                    time.sleep(retry_delay)
                    continue
                else:
                    logger.error(f"File is still locked after {max_retries} attempts. Cannot update.")
                    return 0
            
            today = datetime.now().strftime("%Y%m%d")  # Just the date as string
            
            # Log total items being processed
            total_items_received = len(items)
            logger.info(f"Received {total_items_received} items to update")
            
            # Create a mapping of normalized price_link to item for faster lookup
            item_dict = {}
            for item in items:
                if "price_link" in item and item["price_link"]:
                    # Normalize the price_link to improve matching
                    normalized_link = normalize_url(item["price_link"])
                    item_dict[normalized_link] = item
            
            # Debug: print the number of items we're trying to update
            total_unique_items = len(item_dict)
            if total_unique_items != total_items_received:
                logger.warning(f"Found {total_unique_items} unique items (out of {total_items_received} total)")
            else:
                logger.info(f"All {total_unique_items} items have unique price links")
            
            updated_count = 0
            missing_count = 0
            missing_price_links = []
            
            # For XLSM files, we'll use openpyxl directly to modify only cell values
            if file_ext.lower() == '.xlsm':
                # Load the workbook with read-only initially to check if file is accessible
                try:
                    test_wb = load_workbook(filename=input_file, read_only=True)
                    test_wb.close()
                except Exception as e:
                    if attempt < max_retries:
                        logger.warning(f"Could not open Excel file for reading: {str(e)}. Retry {attempt}/{max_retries}")
                        retry_delay = base_retry_delay * (1 + random.random())
                        time.sleep(retry_delay)
                        continue
                    else:
                        logger.error(f"Could not open Excel file after {max_retries} attempts: {str(e)}")
                        return 0
                
                # Now load the workbook for writing
                try:
                    wb = load_workbook(filename=input_file, keep_vba=True)
                except Exception as e:
                    if attempt < max_retries:
                        logger.warning(f"Could not open Excel file for writing: {str(e)}. Retry {attempt}/{max_retries}")
                        retry_delay = base_retry_delay * (1 + random.random())
                        time.sleep(retry_delay)
                        continue
                    else:
                        logger.error(f"Could not open Excel file after {max_retries} attempts: {str(e)}")
                        return 0
                
                sheet = wb.active  # Assumes data is in the active sheet
                
                # Get column indices
                header_row = 1  # Assuming headers are in row 1
                headers = [cell.value for cell in sheet[header_row]]
                
                # Find column indices (Excel uses 1-based indexing)
                price_link_col = None
                market_player_col = None
                xpath_result_col = None
                stock_status_col = None
                updated_at_col = None
                
                for i, header in enumerate(headers, 1):  # Excel columns are 1-indexed
                    if header == "PriceLink":
                        price_link_col = i
                    elif header == "MarketPlayer":
                        market_player_col = i
                    elif header and header.lower().strip() == "xpath result".lower():
                        xpath_result_col = i
                    elif header and (header.lower().strip() == "stock status".lower() or header.lower().strip() == "out of stock".lower()):
                        stock_status_col = i
                    elif header == "updated_at":
                        updated_at_col = i
                
                # If updated_at column doesn't exist, add it
                if not updated_at_col:
                    updated_at_col = len(headers) + 1
                    sheet.cell(row=header_row, column=updated_at_col).value = "updated_at"
                    logger.info("Added 'updated_at' column to file")
                
                # Log found columns
                logger.info(f"Found columns: PriceLink={price_link_col}, MarketPlayer={market_player_col}, xpath_result={xpath_result_col}, stock_status={stock_status_col}")
                
                # Store original and normalized Excel links for better matching
                excel_links_map = {}
                tracked_items = set()  # Use a set to track which items have been processed
                
                # Count expected rows for this market player
                expected_rows = 0
                if market_player and market_player_col:
                    for row_idx in range(2, sheet.max_row + 1):
                        mp_value = sheet.cell(row=row_idx, column=market_player_col).value
                        if mp_value == market_player:
                            expected_rows += 1
                
                if expected_rows > 0:
                    logger.info(f"Found {expected_rows} rows for market player '{market_player}' in Excel file")
                
                # Update the cells
                for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (after header)
                    if not price_link_col:
                        logger.error("PriceLink column not found in the file!")
                        break
                    
                    # Skip rows that don't match our market player (if we know it)
                    if market_player and market_player_col:
                        mp_value = sheet.cell(row=row_idx, column=market_player_col).value
                        if mp_value != market_player:
                            continue
                    
                    price_link_cell = sheet.cell(row=row_idx, column=price_link_col)
                    price_link = price_link_cell.value
                    
                    if price_link:
                        # Store both original and normalized version
                        original_link = str(price_link)
                        normalized_link = normalize_url(original_link)
                        excel_links_map[normalized_link] = original_link
                        
                        # First try direct match with normalized link
                        if normalized_link in item_dict:
                            item = item_dict[normalized_link]
                            
                            # Update columns
                            if xpath_result_col:
                                sheet.cell(row=row_idx, column=xpath_result_col).value = item["xpath_result"]
                            
                            if stock_status_col:
                                sheet.cell(row=row_idx, column=stock_status_col).value = item["out_of_stock"]
                            
                            # Set updated_at
                            sheet.cell(row=row_idx, column=updated_at_col).value = today
                            updated_count += 1
                            tracked_items.add(normalized_link)
                        else:
                            # Try fuzzy matching by removing parameters or protocol differences
                            found_match = False
                            for item_link in item_dict.keys():
                                # Compare without parameters
                                excel_base = normalized_link.split('?')[0]
                                item_base = item_link.split('?')[0]
                                
                                if excel_base == item_base:
                                    item = item_dict[item_link]
                                    
                                    # Update columns
                                    if xpath_result_col:
                                        sheet.cell(row=row_idx, column=xpath_result_col).value = item["xpath_result"]
                                    
                                    if stock_status_col:
                                        sheet.cell(row=row_idx, column=stock_status_col).value = item["out_of_stock"]
                                    
                                    # Set updated_at
                                    sheet.cell(row=row_idx, column=updated_at_col).value = today
                                    updated_count += 1
                                    tracked_items.add(item_link)
                                    found_match = True
                                    logger.info(f"Matched via fuzzy comparison: {original_link} -> {item['price_link']}")
                                    break
                            
                            if not found_match:
                                # Track missing matches for debugging
                                missing_count += 1
                                if len(missing_price_links) < 10:
                                    missing_price_links.append(original_link)
                
                # Try to save the workbook with retries
                try:
                    # Save the workbook back to the same file
                    wb.save(input_file)
                    logger.info(f"Successfully saved updates to {os.path.basename(input_file)}")
                    break  # Exit the retry loop if successful
                except Exception as e:
                    if attempt < max_retries:
                        logger.warning(f"Could not save Excel file: {str(e)}. Retry {attempt}/{max_retries}")
                        retry_delay = base_retry_delay * (1 + random.random())
                        time.sleep(retry_delay)
                        continue
                    else:
                        logger.error(f"Could not save Excel file after {max_retries} attempts: {str(e)}")
                        return 0
            
            # For non-XLSM files, use pandas approach with improved matching
            else:
                # Try to read the Excel file with retries
                try:
                    df = pd.read_excel(input_file, engine="openpyxl")
                except Exception as e:
                    if attempt < max_retries:
                        logger.warning(f"Could not open Excel file for reading: {str(e)}. Retry {attempt}/{max_retries}")
                        retry_delay = base_retry_delay * (1 + random.random())
                        time.sleep(retry_delay)
                        continue
                    else:
                        logger.error(f"Could not open Excel file after {max_retries} attempts: {str(e)}")
                        return 0
                
                # Find the exact case-sensitive column names, accounting for trailing spaces
                xpath_col = None
                stock_status_col = None
                market_player_col = None
                
                for col in df.columns:
                    if col.lower().strip() == "xpath result".lower():
                        xpath_col = col
                    elif col.lower().strip() == "stock status".lower() or col.lower().strip() == "out of stock".lower():
                        stock_status_col = col
                    elif col == "MarketPlayer":
                        market_player_col = col
                
                # Ensure updated_at column exists with string data type
                if "updated_at" not in df.columns:
                    df["updated_at"] = ""  # Initialize with empty strings, not NaN
                    logger.info("Added 'updated_at' column")
                
                # Log found columns
                logger.info(f"Found columns: XPath={xpath_col}, StockStatus={stock_status_col}, MarketPlayer={market_player_col}")
                
                # Count expected rows for this market player
                expected_rows = 0
                if market_player and market_player_col in df.columns:
                    expected_rows = len(df[df[market_player_col] == market_player])
                    logger.info(f"Found {expected_rows} rows for market player '{market_player}' in Excel file")
                
                # Store original and normalized Excel links
                excel_links_map = {}
                tracked_items = set()  # Use a set to track which items have been processed
                
                # Create a mapping of price_link to row index
                price_link_to_row = {}
                for i, row in df.iterrows():
                    # Skip rows that don't match our market player (if we know it)
                    if market_player and market_player_col in df.columns:
                        if row[market_player_col] != market_player:
                            continue
                    
                    if "PriceLink" in row and row["PriceLink"] is not None:
                        original_link = str(row["PriceLink"])
                        normalized_link = normalize_url(original_link)
                        excel_links_map[normalized_link] = original_link
                        price_link_to_row[normalized_link] = i
                
                # Update rows with improved matching
                for item_link, item in item_dict.items():
                    # First try direct match
                    if item_link in price_link_to_row:
                        i = price_link_to_row[item_link]
                        
                        # Update the exact column names that exist
                        if xpath_col:
                            df.at[i, xpath_col] = item["xpath_result"]
                        
                        if stock_status_col:
                            df.at[i, stock_status_col] = item["out_of_stock"]
                            
                        # Mark as updated today using string value
                        df.at[i, "updated_at"] = today
                        updated_count += 1
                        tracked_items.add(item_link)
                    else:
                        # Try fuzzy matching by removing parameters
                        found_match = False
                        item_base = item_link.split('?')[0]
                        
                        for excel_link, row_idx in price_link_to_row.items():
                            excel_base = excel_link.split('?')[0]
                            
                            if excel_base == item_base:
                                # Update the exact column names that exist
                                if xpath_col:
                                    df.at[row_idx, xpath_col] = item["xpath_result"]
                                
                                if stock_status_col:
                                    df.at[row_idx, stock_status_col] = item["out_of_stock"]
                                    
                                # Mark as updated today using string value
                                df.at[row_idx, "updated_at"] = today
                                updated_count += 1
                                tracked_items.add(item_link)
                                found_match = True
                                logger.info(f"Matched via fuzzy comparison: {excel_links_map[excel_link]} -> {item['price_link']}")
                                break
                        
                        if not found_match:
                            # Track missing matches for debugging
                            missing_count += 1
                            if len(missing_price_links) < 10:
                                missing_price_links.append(item["price_link"])
                
                # Check for items that weren't found in the Excel file
                not_found_items = set(item_dict.keys()) - tracked_items
                if not_found_items:
                    logger.warning(f"Found {len(not_found_items)} items that weren't matched in the Excel file")
                    if len(not_found_items) < 10:
                        logger.warning(f"Unmatched items: {list(not_found_items)}")
                
                # Try to save the dataframe with retries
                try:
                    # Save the dataframe back to the file
                    df.to_excel(input_file, index=False, engine="openpyxl")
                    logger.info(f"Successfully saved updates to {os.path.basename(input_file)}")
                    break  # Exit the retry loop if successful
                except Exception as e:
                    if attempt < max_retries:
                        logger.warning(f"Could not save Excel file: {str(e)}. Retry {attempt}/{max_retries}")
                        retry_delay = base_retry_delay * (1 + random.random())
                        time.sleep(retry_delay)
                        continue
                    else:
                        logger.error(f"Could not save Excel file after {max_retries} attempts: {str(e)}")
                        return 0
            
            # Report statistics and diagnostics
            if missing_count > 0:
                logger.warning(f"⚠️ Failed to find {missing_count} items in the Excel file")
                logger.warning(f"First few missing links: {missing_price_links}")
                
                # Try to find why links might be missing
                if len(missing_price_links) > 0:
                    for missing_link in missing_price_links[:3]:  # Check first 3
                        # Look for partial matches in the Excel file
                        partial_matches = []
                        normalized_missing = normalize_url(missing_link)
                        
                        for excel_link in excel_links_map.keys():
                            # Try various matching approaches
                            if normalized_missing in excel_link or excel_link in normalized_missing:
                                partial_matches.append(excel_links_map[excel_link])
                            # Try matching without domain
                            elif normalized_missing.split('/')[-1] == excel_link.split('/')[-1]:
                                partial_matches.append(excel_links_map[excel_link])
                        
                        if partial_matches:
                            logger.info(f"Partial matches for '{missing_link}': {partial_matches}")
            
            # Final summary
            if expected_rows > 0 and updated_count != expected_rows:
                logger.warning(f"⚠️ Updated {updated_count} rows but expected {expected_rows} rows for {market_player}")
            
            logger.info(f"✅ Updated {updated_count}/{total_unique_items} rows in the file")
            return updated_count
                
        except Exception as e:
            # Handle unexpected exceptions with retry
            if attempt < max_retries:
                logger.warning(f"Error updating file: {str(e)}. Retry {attempt}/{max_retries}")
                retry_delay = base_retry_delay * (1 + random.random())
                time.sleep(retry_delay)
            else:
                logger.error(f"Error updating file after {max_retries} attempts: {str(e)}")
                import traceback
                logger.error(traceback.format_exc())
                return 0
    
    # If we reach here without a return, it means we successfully updated the file on one of the retries
    return updated_count

def filter_rows_not_updated_today(data):
    """
    Filter rows that haven't been updated today.
    If no 'updated_at' column exists, return all rows.
    """
    import logging
    from datetime import datetime
    
    today = datetime.now().strftime("%Y%m%d")
    
    # First, check if updated_at column exists in any row
    has_updated_at = any('updated_at' in row for row in data)
    
    # If no updated_at column, return all rows
    if not has_updated_at:
        logger.info("No 'updated_at' column found. Processing all rows.")
        return data
    
    # Filter only rows that weren't updated today
    filtered_data = []
    for row in data:
        # Get updated_at value, handling possible None or missing
        updated_val = ""
        if 'updated_at' in row and row['updated_at'] is not None:
            updated_val = str(row['updated_at']).strip()
            # Handle possible floating point format (20250413.0)
            updated_val = updated_val.split('.')[0] if '.' in updated_val else updated_val
        
        # Only include rows not updated today
        if updated_val != today:
            filtered_data.append(row)
    
    # Log how many rows were filtered
    skipped = len(data) - len(filtered_data)
    logger.info(f"Filtered out {skipped} rows already updated today. Processing {len(filtered_data)} rows.")
    
    return filtered_data

def get_empty_xpath_results_rows(data):
    """
    Get rows where Xpath Result is empty, regardless of updated_at date.
    Checks multiple possible column names for Xpath Result with improved detection.
    """
    empty_results = []
    possible_column_names = ["Xpath Result", "XPath Result", "xpath result", "XPATH RESULT", "Xpath_Result", "xpath_result"]
    
    # First, find which Xpath Result column name actually exists in the data
    found_column_names = set()
    for row in data[:5]:  # Check first few rows to find column names
        for col in row.keys():
            if "xpath" in col.lower() and "result" in col.lower():
                found_column_names.add(col)
                logger.info(f"Found Xpath Result column: '{col}'")
    
    # Add any found columns to our search list
    for col in found_column_names:
        if col not in possible_column_names:
            possible_column_names.append(col)
    
    logger.info(f"Will check for empty values in these columns: {possible_column_names}")
    
    # Check for empty values in each row
    for row in data:
        is_empty = False
        matched_column = None
        
        # Check all possible column names
        for column_name in possible_column_names:
            if column_name in row:
                value = row[column_name]
                # Check various forms of emptiness
                if value is None:
                    is_empty = True
                    matched_column = column_name
                    break
                elif isinstance(value, str) and value.strip() == "":
                    is_empty = True
                    matched_column = column_name
                    break
                elif isinstance(value, (int, float)) and value == 0:
                    is_empty = True
                    matched_column = column_name
                    break
        
        if is_empty:
            empty_results.append(row)
            # Debug: For the first few empty results, print details
            if len(empty_results) <= 3:
                price_link = row.get("PriceLink", "N/A")
                logger.info(f"Found empty Xpath Result in column '{matched_column}' for URL: {price_link}")
    
    logger.info(f"Found {len(empty_results)} rows with empty Xpath Result fields out of {len(data)} total rows.")
    
    # If no empty results were found, this could indicate a problem
    if len(empty_results) == 0 and len(data) > 0:
        logger.warning("⚠️ No empty Xpath Result rows found. This could indicate an issue with column detection.")
        logger.warning("Here are the column names from the first row for reference:")
        if data:
            logger.warning(str(list(data[0].keys())))
    
    return empty_results